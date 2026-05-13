import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule
from openpyxl.styles.differential import DifferentialStyle
import calendar
from datetime import date, datetime

# ── 색상 팔레트 ──────────────────────────────────────────────
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "BDD7EE"
BLUE_PALE   = "DEEAF1"
ORANGE      = "ED7D31"
ORANGE_PALE = "FCE4D6"
GREEN       = "375623"
GREEN_LIGHT = "E2EFDA"
GRAY_DARK   = "404040"
GRAY_MID    = "A6A6A6"
GRAY_LIGHT  = "F2F2F2"
WHITE       = "FFFFFF"
YELLOW      = "FFD966"
RED         = "C00000"
RED_PALE    = "FFCCCC"

YEAR  = 2026
MONTH = 5

def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium", color=BLUE_MID)
    return Border(left=t, right=t, top=t, bottom=t)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=GRAY_DARK, size=10, name="맑은 고딕"):
    return Font(bold=bold, color=color, size=size, name=name)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# ══════════════════════════════════════════════════════════════
# Sheet 1 – 업무일정
# ══════════════════════════════════════════════════════════════
def build_schedule_sheet(ws):
    ws.title = "업무일정"

    # 열 너비
    col_widths = {
        "A": 6,   # No.
        "B": 13,  # 날짜
        "C": 22,  # 업무명
        "D": 10,  # 담당자
        "E": 12,  # 부서
        "F": 10,  # 우선순위
        "G": 10,  # 진행상태
        "H": 8,   # 진행률%
        "I": 30,  # 비고
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── 헤더 타이틀 ─────────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "📋  업무계획서  |  한국도로공사"
    c.font  = Font(bold=True, color=WHITE, size=15, name="맑은 고딕")
    c.fill  = fill(BLUE_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"작성연월 : {YEAR}년 {MONTH}월"
    c.font  = font(color=WHITE, size=10)
    c.fill  = fill(BLUE_MID)
    c.alignment = center()
    ws.row_dimensions[2].height = 22

    # ── 컬럼 헤더 ───────────────────────────────────────────
    headers = ["No.", "날짜", "업무명", "담당자", "부서",
               "우선순위", "진행상태", "진행률(%)", "비고"]
    header_row = 3
    ws.row_dimensions[header_row].height = 28
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font      = Font(bold=True, color=WHITE, size=10, name="맑은 고딕")
        c.fill      = fill(BLUE_MID)
        c.alignment = center()
        c.border    = thin_border(WHITE)

    # ── 샘플 데이터 ──────────────────────────────────────────
    sample_data = [
        (1,  date(2026,5,4),  "도로 노면 상태 점검",         "김철수", "유지보수팀", "높음", "완료",   100, "4호선 전 구간"),
        (2,  date(2026,5,6),  "교량 안전 정기점검",           "이영희", "안전관리팀", "긴급", "완료",   100, "낙동강교 외 3개소"),
        (3,  date(2026,5,7),  "교통량 데이터 수집·분석",      "박민준", "교통분석팀", "높음", "진행중",  60, "주간보고서 첨부"),
        (4,  date(2026,5,9),  "터널 환기설비 점검",           "최지훈", "시설관리팀", "보통", "진행중",  40, "3개 터널"),
        (5,  date(2026,5,12), "도로 포장 보수공사",           "정소연", "유지보수팀", "높음", "진행중",  75, "경부고속도로 서울-수원"),
        (6,  date(2026,5,13), "월간 업무 보고회의",           "김철수", "전체",       "높음", "예정",     0, "오전 10시, 2층 대회의실"),
        (7,  date(2026,5,14), "휴게소 시설 현황 점검",        "오성민", "시설관리팀", "보통", "예정",     0, "5개소"),
        (8,  date(2026,5,16), "스마트 도로 IoT 센서 점검",    "한미래", "IT혁신팀",   "높음", "예정",     0, "전국 20개 구간"),
        (9,  date(2026,5,19), "고속도로 안전표지 교체",       "이영희", "안전관리팀", "보통", "예정",     0, "노후 표지 120개"),
        (10, date(2026,5,21), "도로 배수시설 점검",           "박민준", "유지보수팀", "보통", "예정",     0, "집중호우 대비"),
        (11, date(2026,5,23), "교통사고 분석 보고서 제출",    "정소연", "교통분석팀", "높음", "예정",     0, "4월 사고 데이터"),
        (12, date(2026,5,26), "도로 조명 일제 점검",          "오성민", "시설관리팀", "보통", "예정",     0, "야간 점검"),
        (13, date(2026,5,28), "차세대 톨링시스템 테스트",     "한미래", "IT혁신팀",   "긴급", "예정",     0, "시범구간 3곳"),
        (14, date(2026,5,30), "5월 업무 마감 및 결과보고",    "김철수", "전체",       "높음", "예정",     0, "팀장급 취합"),
    ]

    priority_color = {"긴급": RED_PALE, "높음": ORANGE_PALE, "보통": BLUE_PALE, "낮음": GRAY_LIGHT}
    status_color   = {"완료": GREEN_LIGHT, "진행중": YELLOW+"80" if False else "FFF2CC", "예정": GRAY_LIGHT}

    for r_idx, row in enumerate(sample_data, start=4):
        ws.row_dimensions[r_idx].height = 22
        no, dt, task, person, dept, priority, status, pct, note = row
        values = [no, dt.strftime("%Y-%m-%d"), task, person, dept, priority, status, pct, note]

        for c_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border    = thin_border()
            c.font      = font(size=10)
            c.alignment = center() if c_idx not in (3, 9) else left(wrap=True)

            # 행 배경
            base_fill = GRAY_LIGHT if r_idx % 2 == 0 else WHITE
            c.fill = fill(base_fill)

        # 우선순위 셀 색상
        ws.cell(row=r_idx, column=6).fill = fill(priority_color.get(priority, WHITE))
        ws.cell(row=r_idx, column=6).font = Font(bold=True, size=10, name="맑은 고딕",
            color=RED if priority=="긴급" else GRAY_DARK)

        # 상태 셀 색상
        ws.cell(row=r_idx, column=7).fill = fill(status_color.get(status, WHITE))

        # 날짜 포맷
        ws.cell(row=r_idx, column=2).number_format = "YYYY-MM-DD"

        # 진행률 색상
        pct_cell = ws.cell(row=r_idx, column=8)
        if pct == 100:
            pct_cell.fill = fill("C6EFCE")
            pct_cell.font = Font(bold=True, color="375623", size=10, name="맑은 고딕")
        elif pct >= 50:
            pct_cell.fill = fill("FFEB9C")
        else:
            pct_cell.fill = fill(base_fill)

    # ── 범례 ────────────────────────────────────────────────
    legend_row = 4 + len(sample_data) + 1
    ws.merge_cells(f"A{legend_row}:I{legend_row}")
    c = ws[f"A{legend_row}"]
    c.value = "【 범례 】  우선순위: 긴급(빨강) / 높음(주황) / 보통(파랑) / 낮음(회색)    진행상태: 완료(초록) / 진행중(노랑) / 예정(회색)"
    c.font  = font(size=9, color=GRAY_MID)
    c.alignment = left()
    c.fill  = fill(GRAY_LIGHT)
    ws.row_dimensions[legend_row].height = 20

    # ── 외곽 두꺼운 테두리 ──────────────────────────────────
    last_data_row = 3 + len(sample_data)
    for r in range(1, last_data_row + 1):
        for col in range(1, 10):
            c = ws.cell(row=r, column=col)
            existing = c.border
            left_s   = Side(style="medium", color=BLUE_MID) if col == 1 else existing.left
            right_s  = Side(style="medium", color=BLUE_MID) if col == 9 else existing.right
            top_s    = Side(style="medium", color=BLUE_MID) if r == 1 else existing.top
            bottom_s = Side(style="medium", color=BLUE_MID) if r == last_data_row else existing.bottom
            c.border = Border(left=left_s, right=right_s, top=top_s, bottom=bottom_s)

    # 틀 고정 (헤더 3행 고정)
    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════
# Sheet 2 – 달력
# ══════════════════════════════════════════════════════════════
def build_calendar_sheet(ws, sample_data_for_calendar):
    ws.title = "달력"

    # 달력 계산
    cal = calendar.monthcalendar(YEAR, MONTH)
    month_name = f"{YEAR}년 {MONTH}월"

    # 열 너비: 7열 (일~토)
    day_col_width = 16
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = day_col_width

    # ── 제목 ────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"📅  {month_name}  업무 달력  |  한국도로공사"
    c.font  = Font(bold=True, color=WHITE, size=15, name="맑은 고딕")
    c.fill  = fill(BLUE_DARK)
    c.alignment = center()
    ws.row_dimensions[1].height = 38

    # 안내 문구
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "※  업무일정 시트에 데이터 입력 시 달력이 자동으로 업데이트됩니다."
    c.font  = Font(italic=True, color=BLUE_MID, size=9, name="맑은 고딕")
    c.fill  = fill(BLUE_PALE)
    c.alignment = center()
    ws.row_dimensions[2].height = 18

    # ── 요일 헤더 ───────────────────────────────────────────
    day_names = ["일", "월", "화", "수", "목", "금", "토"]
    day_colors = [RED, BLUE_DARK, BLUE_DARK, BLUE_DARK, BLUE_DARK, BLUE_DARK, "1F497D"]
    for col_idx, (day, color) in enumerate(zip(day_names, day_colors), start=1):
        c = ws.cell(row=3, column=col_idx, value=day)
        c.font      = Font(bold=True, color=WHITE, size=12, name="맑은 고딕")
        c.fill      = fill(color)
        c.alignment = center()
        c.border    = thin_border(WHITE)
    ws.row_dimensions[3].height = 28

    # ── 날짜별 업무 매핑 ────────────────────────────────────
    tasks_by_date = {}
    for row in sample_data_for_calendar:
        _, dt, task, person, _, priority, status, _, _ = row
        key = dt.day
        if key not in tasks_by_date:
            tasks_by_date[key] = []
        tasks_by_date[key].append((task, person, priority, status))

    # ── 주(週) 행 렌더링 ─────────────────────────────────────
    today = date.today()
    row_start = 4

    for week_idx, week in enumerate(cal):
        # 날짜 번호 행
        num_row = row_start + week_idx * 4
        # 내용 행 3개
        task_rows = [num_row + 1, num_row + 2, num_row + 3]
        ws.row_dimensions[num_row].height = 22
        for tr in task_rows:
            ws.row_dimensions[tr].height = 18

        for col_idx, day_num in enumerate(week, start=1):
            # 날짜 번호 셀
            num_cell = ws.cell(row=num_row, column=col_idx)
            is_sunday   = col_idx == 1
            is_saturday = col_idx == 7
            is_today    = (day_num != 0 and date(YEAR, MONTH, day_num) == today)

            if day_num == 0:
                num_cell.fill   = fill(GRAY_LIGHT)
                bg = GRAY_LIGHT
            elif is_today:
                num_cell.value  = f"▶ {day_num}"
                num_cell.fill   = fill(ORANGE)
                num_cell.font   = Font(bold=True, color=WHITE, size=11, name="맑은 고딕")
                bg = ORANGE_PALE
            else:
                num_cell.value  = day_num
                bg_day          = "FFF5F5" if is_sunday else ("F0F4FF" if is_saturday else WHITE)
                bg              = bg_day
                num_cell.fill   = fill(bg_day)
                day_color       = RED if is_sunday else ("1F497D" if is_saturday else GRAY_DARK)
                num_cell.font   = Font(bold=True, color=day_color, size=11, name="맑은 고딕")

            num_cell.alignment = center()
            num_cell.border    = thin_border()

            # 업무 내용 채우기
            tasks = tasks_by_date.get(day_num, []) if day_num != 0 else []
            for slot_idx, task_row_num in enumerate(task_rows):
                tc = ws.cell(row=task_row_num, column=col_idx)
                tc.fill   = fill(GRAY_LIGHT if day_num == 0 else bg)
                tc.border = thin_border()
                if slot_idx < len(tasks):
                    task_name, person, priority, status = tasks[slot_idx]
                    display = f"• {task_name[:10]}…" if len(task_name) > 10 else f"• {task_name}"
                    tc.value     = display
                    tc.font      = Font(size=8, name="맑은 고딕",
                                        color=RED if priority=="긴급" else
                                              ORANGE if priority=="높음" else GRAY_DARK)
                    tc.alignment = left(wrap=False)
                    # 상태별 배경
                    if status == "완료":
                        tc.fill = fill("E2EFDA")
                    elif status == "진행중":
                        tc.fill = fill("FFF2CC")

    # ── 범례 박스 ────────────────────────────────────────────
    legend_row = row_start + len(cal) * 4 + 1
    ws.merge_cells(f"A{legend_row}:G{legend_row}")
    c = ws[f"A{legend_row}"]
    c.value = ("【 범례 】  "
               "배경: 완료=초록 / 진행중=노랑 / 예정=흰색    "
               "글씨: 긴급=빨강 / 높음=주황 / 보통=검정    "
               "▶ = 오늘")
    c.font  = Font(italic=True, size=8, color=GRAY_MID, name="맑은 고딕")
    c.fill  = fill(GRAY_LIGHT)
    c.alignment = left()
    ws.row_dimensions[legend_row].height = 18

    # 틀 고정
    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    ws1 = wb.active

    sample_data = [
        (1,  date(2026,5,4),  "도로 노면 상태 점검",         "김철수", "유지보수팀", "높음", "완료",   100, "4호선 전 구간"),
        (2,  date(2026,5,6),  "교량 안전 정기점검",           "이영희", "안전관리팀", "긴급", "완료",   100, "낙동강교 외 3개소"),
        (3,  date(2026,5,7),  "교통량 데이터 수집·분석",      "박민준", "교통분석팀", "높음", "진행중",  60, "주간보고서 첨부"),
        (4,  date(2026,5,9),  "터널 환기설비 점검",           "최지훈", "시설관리팀", "보통", "진행중",  40, "3개 터널"),
        (5,  date(2026,5,12), "도로 포장 보수공사",           "정소연", "유지보수팀", "높음", "진행중",  75, "경부고속도로 서울-수원"),
        (6,  date(2026,5,13), "월간 업무 보고회의",           "김철수", "전체",       "높음", "예정",     0, "오전 10시, 2층 대회의실"),
        (7,  date(2026,5,14), "휴게소 시설 현황 점검",        "오성민", "시설관리팀", "보통", "예정",     0, "5개소"),
        (8,  date(2026,5,16), "스마트 도로 IoT 센서 점검",    "한미래", "IT혁신팀",   "높음", "예정",     0, "전국 20개 구간"),
        (9,  date(2026,5,19), "고속도로 안전표지 교체",       "이영희", "안전관리팀", "보통", "예정",     0, "노후 표지 120개"),
        (10, date(2026,5,21), "도로 배수시설 점검",           "박민준", "유지보수팀", "보통", "예정",     0, "집중호우 대비"),
        (11, date(2026,5,23), "교통사고 분석 보고서 제출",    "정소연", "교통분석팀", "높음", "예정",     0, "4월 사고 데이터"),
        (12, date(2026,5,26), "도로 조명 일제 점검",          "오성민", "시설관리팀", "보통", "예정",     0, "야간 점검"),
        (13, date(2026,5,28), "차세대 톨링시스템 테스트",     "한미래", "IT혁신팀",   "긴급", "예정",     0, "시범구간 3곳"),
        (14, date(2026,5,30), "5월 업무 마감 및 결과보고",    "김철수", "전체",       "높음", "예정",     0, "팀장급 취합"),
    ]

    build_schedule_sheet(ws1)
    ws2 = wb.create_sheet()
    build_calendar_sheet(ws2, sample_data)

    # 시트 탭 색상
    ws1.sheet_properties.tabColor = "2E75B6"
    ws2.sheet_properties.tabColor = "ED7D31"

    out_path = r"C:\Users\L2404161\Desktop\도로공사_바이브코딩\업무계획서\업무계획서_2026년5월.xlsx"
    wb.save(out_path)
    print(f"저장 완료: {out_path}")

if __name__ == "__main__":
    main()
